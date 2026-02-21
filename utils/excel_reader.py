from openpyxl import load_workbook

def read_excel(path, sheet=None):
    wb = load_workbook(path)

    # If sheet not given, take first sheet automatically
    if sheet:
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data